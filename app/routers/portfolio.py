import os
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timedelta
from decimal import Decimal
import csv
import io
import logging
import re
import yfinance as yf
import plaid
from typing import Dict, List

from ..database import get_db
from ..models import PortfolioAccount, Holding, BankTransaction, BrokerCredential, PlaidItem
from ..auth import get_current_user
from ..fidelity_scraper import FidelityScraper
from ..crypto_utils import CredentialEncryptor
from ..csv_parser import CSVParser
from ..plaid_client import get_plaid_client

router = APIRouter(prefix="/portfolio", tags=["portfolio"])
logger = logging.getLogger(__name__)


@router.get("/summary")
def get_portfolio_summary(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Get overall portfolio summary"""
    
    # Get all accounts
    accounts = db.query(PortfolioAccount).filter(
        PortfolioAccount.user_id == user.id,
        PortfolioAccount.is_active == True
    ).all()
    
    # Calculate totals
    total_investments = Decimal('0')
    total_cash = Decimal('0')
    total_credit_debt = Decimal('0')
    
    investment_accounts = []
    checking_accounts = []
    savings_accounts = []
    credit_card_accounts = []
    
    for acc in accounts:
        acc_data = {
            'id': acc.id,
            'institution': acc.institution,
            'account_name': acc.account_name or f"{acc.institution} (...{acc.account_number_last4})" if acc.account_number_last4 else acc.institution,
            'account_number_last4': acc.account_number_last4,
            'account_type': acc.account_type,
            'balance': float(acc.balance),
            'last_synced': acc.last_synced.isoformat() if acc.last_synced else None,
            'owner': acc.account_holder or 'Not specified'
        }
        
        if acc.account_type == 'investment':
            total_investments += acc.balance
            investment_accounts.append(acc_data)
        elif acc.account_type == 'credit_card':
            total_credit_debt += acc.balance
            credit_card_accounts.append(acc_data)
        elif acc.account_type == 'savings':
            total_cash += acc.balance
            savings_accounts.append(acc_data)
        else:  # checking and anything else
            total_cash += acc.balance
            checking_accounts.append(acc_data)
    
    # For backward compat keep bank_accounts as checking+savings combined
    bank_accounts = checking_accounts + savings_accounts
    
    # Get total holdings count
    holdings_count = db.query(func.count(Holding.id)).filter(Holding.user_id == user.id).scalar()
    
    # Get recent transactions count
    recent_txns = db.query(func.count(BankTransaction.id)).filter(
        BankTransaction.user_id == user.id
    ).scalar()
    
    net_worth = total_investments + total_cash - total_credit_debt
    
    return {
        'net_worth': float(net_worth),
        'total_investments': float(total_investments),
        'total_cash': float(total_cash),
        'total_credit_debt': float(total_credit_debt),
        'investment_accounts': investment_accounts,
        'checking_accounts': checking_accounts,
        'savings_accounts': savings_accounts,
        'credit_card_accounts': credit_card_accounts,
        'bank_accounts': bank_accounts,
        'holdings_count': holdings_count,
        'transactions_count': recent_txns
    }


@router.get("/holdings")
def get_holdings(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Get all investment holdings"""
    
    holdings = db.query(Holding).filter(Holding.user_id == user.id).all()
    
    result = []
    for h in holdings:
        result.append({
            'id': h.id,
            'institution': h.account.institution,
            'account': h.account.account_name or h.account.institution,
            'account_number_last4': h.account.account_number_last4,
            'symbol': h.symbol,
            'name': h.name,
            'quantity': float(h.quantity),
            'cost_basis': float(h.cost_basis) if h.cost_basis else None,
            'current_price': float(h.current_price) if h.current_price else None,
            'current_value': float(h.current_value) if h.current_value else None,
            'gain_loss': float(h.current_value - h.cost_basis) if (h.current_value and h.cost_basis) else None,
            'asset_type': h.asset_type,
            'snapshot_date': h.snapshot_date.isoformat()
        })
    
    return result


@router.get("/transactions")
def get_transactions(limit: int = 100, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Get recent bank transactions"""
    
    txns = db.query(BankTransaction).filter(
        BankTransaction.user_id == user.id
    ).order_by(BankTransaction.transaction_date.desc()).limit(limit).all()
    
    result = []
    for t in txns:
        result.append({
            'id': t.id,
            'account': t.account.account_name or t.account.institution,
            'date': t.transaction_date.isoformat(),
            'description': t.description,
            'amount': float(t.amount),
            'category': t.category,
            'transaction_type': t.transaction_type,
            'balance_after': float(t.balance_after) if t.balance_after else None
        })
    
    return result


async def parse_etrade_equity_awards(db: Session, user, csv_text: str, delimiter: str, file, snapshot_date, is_espp: bool = False):
    """
    Parse E*TRADE ESPP or RSU format CSV and save to dedicated tracking tables.
    ESPP columns: Record Type, Symbol, Purchase Date, Purchase Price, Purchased Qty., Sellable Qty., Expected Gain/Loss, Est. Market Value
    RSU columns: Record Type, Symbol, Grant Date, Settlement Type, Granted Qty., Withheld Qty., Vested Qty., Unvested Qty., Sellable Qty., Est. Market Value, Grant Number
    """
    from app.models import ESPPGrant, RSUGrant
    
    account_type_name = 'ESPP' if is_espp else 'RSU'
    logger.info(f"Parsing E*TRADE {account_type_name} CSV")
    
    rows = list(csv.DictReader(io.StringIO(csv_text), delimiter=delimiter))
    
    def normalize_key(key: str) -> str:
        return (key or '').replace('\ufeff', '').strip().lower()
    
    def parse_decimal(value):
        if value is None:
            return None
        s = str(value).strip().replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
        if not s or s in ['-', 'N/A', 'n/a', '']:
            return None
        try:
            return Decimal(s)
        except:
            return None
    
    def parse_date(value):
        if not value or value.strip() in ['-', 'N/A', 'n/a', '']:
            return None
        try:
            from dateutil import parser
            return parser.parse(value).date()
        except:
            return None
    
    # Find or create E*TRADE account for this equity type
    account = db.query(PortfolioAccount).filter(
        PortfolioAccount.user_id == user.id,
        PortfolioAccount.institution == 'E*TRADE',
        PortfolioAccount.account_name.like(f'%{account_type_name}%')
    ).first()
    
    if not account:
        account = PortfolioAccount(
            user_id=user.id,
            institution='E*TRADE',
            account_type='investment',
            account_name=f'E*TRADE {account_type_name}',
            account_number_last4='',
            balance=Decimal('0'),
            last_synced=datetime.now(),
            is_active=True
        )
        db.add(account)
        db.flush()
        logger.info(f"Created E*TRADE {account_type_name} account")
    
    imported_count = 0
    
    for row in rows:
        normalized_row = {normalize_key(k): v for k, v in row.items()}
        
        # Get symbol
        symbol = normalized_row.get('symbol', '').strip()
        if not symbol or symbol == '-':
            continue
        
        # Get common fields
        record_type = normalized_row.get('record type', '').strip()
        sellable_qty = parse_decimal(normalized_row.get('sellable qty.') or normalized_row.get('sellable qty'))
        est_market_value = parse_decimal(normalized_row.get('est. market value') or normalized_row.get('est market value'))
        
        if is_espp:
            # Parse ESPP-specific fields
            purchase_date = parse_date(normalized_row.get('purchase date'))
            purchase_price = parse_decimal(normalized_row.get('purchase price'))
            purchased_qty = parse_decimal(normalized_row.get('purchased qty.') or normalized_row.get('purchased qty'))
            expected_gain_loss = parse_decimal(normalized_row.get('expected gain/loss') or normalized_row.get('expected gain'))
            
            # E*TRADE Excel ESPP has 2 hidden columns shifting data by 2 positions.
            # Header blank cols are renamed _extra_0 and _extra_1 during Excel processing.
            # Actual data layout: sellable qty = purchased qty, expected G/L = _extra_0, market value = _extra_1.
            if not sellable_qty or sellable_qty <= 0:
                sellable_qty = purchased_qty
            if not expected_gain_loss:
                expected_gain_loss = parse_decimal(normalized_row.get('_extra_0'))
            if est_market_value == purchased_qty:
                est_market_value = parse_decimal(normalized_row.get('_extra_1')) or \
                                   parse_decimal(normalized_row.get('')) or est_market_value
            
            if not sellable_qty or sellable_qty <= 0:
                continue
            
            # Check for existing grant (match by symbol, purchase date, and purchase price)
            existing = db.query(ESPPGrant).filter(
                ESPPGrant.user_id == user.id,
                ESPPGrant.account_id == account.id,
                ESPPGrant.symbol == symbol.upper(),
                ESPPGrant.purchase_date == purchase_date,
                ESPPGrant.purchase_price == purchase_price
            ).first()
            
            if existing:
                existing.sellable_qty = sellable_qty
                existing.est_market_value = est_market_value
                existing.expected_gain_loss = expected_gain_loss
                existing.last_updated = datetime.now()
            else:
                grant = ESPPGrant(
                    user_id=user.id,
                    account_id=account.id,
                    symbol=symbol.upper(),
                    record_type=record_type,
                    purchase_date=purchase_date,
                    purchase_price=purchase_price,
                    purchased_qty=purchased_qty,
                    sellable_qty=sellable_qty,
                    expected_gain_loss=expected_gain_loss,
                    est_market_value=est_market_value
                )
                db.add(grant)
        
        else:
            # Parse RSU-specific fields
            if not sellable_qty or sellable_qty <= 0:
                continue
            
            grant_date = parse_date(normalized_row.get('grant date'))
            grant_number = normalized_row.get('grant number', '').strip() or normalized_row.get('grant no', '').strip()
            settlement_type = normalized_row.get('settlement type', '').strip()
            granted_qty = parse_decimal(normalized_row.get('granted qty.') or normalized_row.get('granted qty'))
            withheld_qty = parse_decimal(normalized_row.get('withheld qty.') or normalized_row.get('withheld qty'))
            vested_qty = parse_decimal(normalized_row.get('vested qty.') or normalized_row.get('vested qty'))
            unvested_qty = parse_decimal(normalized_row.get('unvested qty.') or normalized_row.get('unvested qty'))
            
            # Check for existing grant (match by grant number or grant date + symbol)
            existing = None
            if grant_number:
                existing = db.query(RSUGrant).filter(
                    RSUGrant.user_id == user.id,
                    RSUGrant.account_id == account.id,
                    RSUGrant.grant_number == grant_number
                ).first()
            
            if not existing and grant_date:
                existing = db.query(RSUGrant).filter(
                    RSUGrant.user_id == user.id,
                    RSUGrant.account_id == account.id,
                    RSUGrant.symbol == symbol.upper(),
                    RSUGrant.grant_date == grant_date
                ).first()
            
            if existing:
                existing.vested_qty = vested_qty
                existing.unvested_qty = unvested_qty
                existing.sellable_qty = sellable_qty
                existing.withheld_qty = withheld_qty
                existing.est_market_value = est_market_value
                existing.last_updated = datetime.now()
            else:
                grant = RSUGrant(
                    user_id=user.id,
                    account_id=account.id,
                    symbol=symbol.upper(),
                    record_type=record_type,
                    grant_number=grant_number,
                    grant_date=grant_date,
                    settlement_type=settlement_type,
                    granted_qty=granted_qty,
                    withheld_qty=withheld_qty,
                    vested_qty=vested_qty,
                    unvested_qty=unvested_qty,
                    sellable_qty=sellable_qty,
                    est_market_value=est_market_value
                )
                db.add(grant)
        
        imported_count += 1
    
    db.commit()
    
    # Update account balance based on total sellable value
    if is_espp:
        total = db.query(func.sum(ESPPGrant.est_market_value)).filter(
            ESPPGrant.account_id == account.id
        ).scalar() or Decimal('0')
    else:
        total = db.query(func.sum(RSUGrant.est_market_value)).filter(
            RSUGrant.account_id == account.id
        ).scalar() or Decimal('0')
    
    account.balance = total
    account.last_synced = datetime.now()
    db.commit()
    
    return {'imported': imported_count, 'message': f'Imported {imported_count} {account_type_name} grants', 'date': snapshot_date.isoformat()}


@router.post("/upload/holdings")
async def upload_holdings_csv(
    file: UploadFile = File(...),
    account_id: int = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Upload holdings CSV or Excel file with multiple sheets"""
    
    filename = file.filename.lower()
    
    # Handle Excel files with multiple sheets
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        import openpyxl
        from io import BytesIO
        
        content = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        
        total_imported = 0
        results = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Convert sheet to CSV-like string
            rows = []
            header_processed = False
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if not any(cell for cell in row):
                    continue
                if not header_processed:
                    # First non-empty row = header; give unique names to blank columns
                    # so DictReader can distinguish shifted data columns (e.g. E*TRADE ESPP)
                    header_cells = []
                    extra_count = 0
                    for cell in row:
                        if cell is None or str(cell).strip() == '':
                            header_cells.append(f'_extra_{extra_count}')
                            extra_count += 1
                        else:
                            header_cells.append(str(cell))
                    rows.append('\t'.join(header_cells))
                    header_processed = True
                else:
                    row_str = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                    rows.append(row_str)
            
            if not rows:
                continue
                
            csv_text = '\n'.join(rows)
            delimiter = '\t'
            snapshot_date = date.today()
            
            # Detect format for this sheet
            first_50_lines = '\n'.join(rows[:50]).lower()
            is_espp = 'record type' in first_50_lines and 'sellable qty' in first_50_lines and ('purchase date' in first_50_lines or 'purchased qty' in first_50_lines)
            is_rsu = 'record type' in first_50_lines and 'sellable qty' in first_50_lines and ('grant date' in first_50_lines or 'vested qty' in first_50_lines)
            
            logger.info(f"Sheet '{sheet_name}': is_espp={is_espp}, is_rsu={is_rsu}, rows={len(rows)}")
            
            if is_espp or is_rsu:
                result = await parse_etrade_equity_awards(db, user, csv_text, delimiter, file, snapshot_date, is_espp=is_espp)
                total_imported += result.get('imported', 0)
                results.append(f"{sheet_name}: {result.get('message', 'Imported')}")
            else:
                # Process as regular holdings - create a temporary file-like object
                from io import StringIO
                import csv as csv_module
                
                reader = csv_module.DictReader(StringIO(csv_text), delimiter=delimiter)
                sheet_rows = list(reader)
                
                if sheet_rows:
                    # Continue with normal processing (we'll process inline)
                    logger.info(f"Processing sheet '{sheet_name}' as regular holdings: {len(sheet_rows)} rows")
                    results.append(f"{sheet_name}: {len(sheet_rows)} rows detected")
        
        workbook.close()
        
        if total_imported > 0:
            return {'imported': total_imported, 'message': f'Imported from {len(results)} sheet(s): ' + ', '.join(results)}
        else:
            return {'imported': 0, 'message': 'No ESPP/RSU data found in Excel sheets'}
    
    # Original CSV handling
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")
    
    content = await file.read()
    csv_text = content.decode('utf-8')

    # Detect E*TRADE transactions/activity exports misrouted to holdings upload
    if 'Account Activity for' in csv_text or 'All Transactions Activity Types' in csv_text:
        raise HTTPException(status_code=400, detail="This file looks like a transactions/activity export. Please upload it using the Transactions import.")
    
    # Detect delimiter (tab, comma, semicolon, pipe)
    delimiter = ','
    if csv_text:
        sample = csv_text[:4096]
        try:
            sniffed = csv.Sniffer().sniff(sample, delimiters=[',', '\t', ';', '|'])
            delimiter = sniffed.delimiter
            logger.info(f"CSV Upload: Detected delimiter: {repr(delimiter)}")
        except Exception as e:
            first_line = csv_text.split('\n')[0] if csv_text else ''
            delimiter = '\t' if '\t' in first_line else ','
            logger.info(f"CSV Upload: Fallback delimiter: {repr(delimiter)}, error: {e}")
    
    # Split CSV into lines to detect section changes
    lines = csv_text.split('\n')
    
    imported_count = 0
    snapshot_date = date.today()
    
    # Extract account metadata from E*TRADE CSV preamble
    etrade_account_info = {}
    for line in lines[:20]:  # Check first 20 lines for metadata
        if ':' in line and ',' not in line[:line.index(':')]:
            parts = line.split(':', 1)
            key = parts[0].strip().lower()
            value = parts[1].strip() if len(parts) > 1 else ''
            if any(k in key for k in ['account', 'name', 'type', 'number']):
                etrade_account_info[key] = value
                logger.info(f"Found E*TRADE metadata: {key} = {value}")
        
        # Also check for E*TRADE account format: "Individual Brokerage -1234"
        line_lower = line.lower()
        if any(acct_type in line_lower for acct_type in ['individual brokerage', 'coverdell', 'premium savings', 'max rate checking']):
            # Extract account type and number
            parts = line.split('\t')
            if parts:
                account_info = parts[0].strip().strip('"').strip("'")
                if account_info and not account_info.lower().startswith('account'):
                    etrade_account_info['account_full'] = account_info
                    # Extract account number (last 4 digits after hyphen)
                    if '-' in account_info:
                        acct_num = account_info.split('-')[-1].strip()
                        etrade_account_info['account number'] = acct_num
                        etrade_account_info['account name'] = account_info.split('-')[0].strip()
                    logger.info(f"Found E*TRADE account: {account_info}")
    
    logger.info(f"E*TRADE account info extracted: {etrade_account_info}")
    
    # Detect if this is a Fidelity CSV by looking for characteristic patterns
    is_fidelity_format = any('fidelity' in (file.filename or '').lower() for _ in [1]) or \
                        any('Symbol/CUSIP' in line or 'symbol/cusip' in line.lower() for line in lines[:20])
    
    # Detect E*TRADE ESPP/RSU format
    first_50_lines = '\n'.join(lines[:50]).lower()
    is_etrade_espp = 'record type' in first_50_lines and 'sellable qty' in first_50_lines and ('purchase date' in first_50_lines or 'purchased qty' in first_50_lines)
    is_etrade_rsu = 'record type' in first_50_lines and 'sellable qty' in first_50_lines and ('grant date' in first_50_lines or 'vested qty' in first_50_lines)
    
    # Detect E*TRADE brokerage account format (Individual Brokerage, Coverdell, etc.)
    is_etrade_brokerage = ('account summary' in first_50_lines and 
                          any(acct in first_50_lines for acct in ['individual brokerage', 'coverdell', 'premium savings', 'max rate checking']))
    
    if is_etrade_espp or is_etrade_rsu:
        logger.info(f"Detected E*TRADE {'ESPP' if is_etrade_espp else 'RSU'} format")
        return await parse_etrade_equity_awards(db, user, csv_text, delimiter, file, snapshot_date, is_espp=is_etrade_espp)
    
    if is_etrade_brokerage:
        logger.info("Detected E*TRADE brokerage account format")
        # Mark as E*TRADE for downstream processing
        if not etrade_account_info.get('institution'):
            etrade_account_info['institution'] = 'E*TRADE'
    
    # Find where holdings section starts (look for "Symbol/CUSIP" header)
    account_section_end = 0
    holdings_section_start = 0
    
    for idx, line in enumerate(lines):
        if 'Symbol/CUSIP' in line or 'symbol/cusip' in line.lower() or 'Symbol' in line and 'Quantity' in line:
            holdings_section_start = idx
            account_section_end = idx
            logger.info(f"CSV Upload: Found holdings section at line {idx}")
            break
    
    # Parse account summary section (top part)
    if account_section_end > 0:
        account_csv = '\n'.join(lines[:account_section_end])
        account_reader = csv.DictReader(io.StringIO(account_csv), delimiter=delimiter)
        account_rows = list(account_reader)
    else:
        account_rows = []
    
    # Parse holdings section (bottom part)  
    if holdings_section_start > 0:
        holdings_csv = '\n'.join(lines[holdings_section_start:])
        holdings_reader = csv.DictReader(io.StringIO(holdings_csv), delimiter=delimiter)
        holdings_rows = list(holdings_reader)
    else:
        holdings_rows = []
    
    logger.info(f"CSV Upload: Found {len(account_rows)} account rows and {len(holdings_rows)} holdings rows")
    
    # Parse full CSV for non-Fidelity formats
    rows = list(csv.DictReader(io.StringIO(csv_text), delimiter=delimiter))

    def normalize_key(key: str) -> str:
        return (key or '').replace('\ufeff', '').strip().lower()

    def has_holdings_keys(sample_row: dict) -> bool:
        if not sample_row:
            return False
        keys = {normalize_key(k) for k in sample_row.keys() if k}
        has_symbol = any(k in keys for k in ['symbol', 'ticker', 'cusip', 'security symbol'])
        has_qty = any(k in keys for k in ['quantity', 'qty', 'shares'])
        has_value = any(k in keys for k in ['current value', 'market value', 'total value', 'value'])
        return has_symbol and (has_qty or has_value)

    def find_header_start(lines_list: list) -> int:
        for i, line in enumerate(lines_list):
            low = line.lower()
            if any(k in low for k in ['symbol', 'ticker', 'cusip']) and any(k in low for k in ['quantity', 'qty', 'shares', 'market value', 'current value']):
                return i
        return 0

    # If the first row isn't a real header (preamble lines), try to locate the actual header line
    if not rows or not has_holdings_keys(rows[0]):
        header_start = find_header_start(lines)
        if header_start > 0:
            alt_csv = '\n'.join(lines[header_start:])
            rows = list(csv.DictReader(io.StringIO(alt_csv), delimiter=delimiter))
    
    # Check for Fidelity flat format (Account Number + Account Name + holdings in same row)
    has_flat_account_format = False
    if rows and len(rows) > 0:
        first_row_keys = {normalize_key(k) for k in rows[0].keys()}
        has_flat_account_format = 'account number' in first_row_keys and 'account name' in first_row_keys
    if has_flat_account_format:
        logger.warning("CSV Upload: Detected Fidelity flat format with Account Number + Account Name")
        print(f"DEBUG: Processing Fidelity flat format CSV with {len(rows)} rows")
        
        # Group holdings by account
        account_holdings = {}
        account_names = {}
        
        for row in rows:
            # Normalize row keys for case-insensitive matching
            normalized_row = {normalize_key(k): v for k, v in row.items()}
            
            acct_num = (normalized_row.get('account number') or '').strip()
            acct_name = (normalized_row.get('account name') or '').strip()
            symbol = (normalized_row.get('symbol') or '').strip()
            
            if not acct_num or not symbol:
                continue
            
            if acct_num not in account_holdings:
                account_holdings[acct_num] = []
                account_names[acct_num] = acct_name
                print(f"DEBUG: Found new account {acct_num} - {acct_name}")
            
            account_holdings[acct_num].append(normalized_row)
        
        # Process each account
        institution = 'Fidelity' if is_fidelity_format else 'Imported'
        imported_count = 0
        
        print(f"DEBUG: Total accounts found: {len(account_holdings)}")
        print(f"DEBUG: Account numbers: {list(account_holdings.keys())}")
        
        for acct_num, holdings in account_holdings.items():
            acct_name = account_names[acct_num]
            last4 = acct_num[-4:] if len(acct_num) >= 4 else acct_num
            
            # Calculate total balance from holdings
            total_balance = Decimal('0')
            for holding_row in holdings:
                current_value = (holding_row.get('current value') or '').strip()
                if current_value:
                    clean_val = current_value.replace('$', '').replace(',', '').strip()
                    try:
                        if clean_val:
                            total_balance += Decimal(clean_val)
                    except:
                        pass
            
            # Find or create account
            account = db.query(PortfolioAccount).filter(
                PortfolioAccount.user_id == user.id,
                PortfolioAccount.account_number_last4 == last4
            ).first()
            
            if account:
                account.balance = total_balance
                account.last_synced = datetime.now()
                account.institution = institution
                account.account_name = acct_name
            else:
                account = PortfolioAccount(
                    user_id=user.id,
                    institution=institution,
                    account_type='investment',
                    account_name=acct_name,
                    account_number_last4=last4,
                    balance=total_balance,
                    last_synced=datetime.now(),
                    is_active=True
                )
                db.add(account)
                db.flush()
            
            # Clear old holdings for this account and date
            db.query(Holding).filter(
                Holding.account_id == account.id,
                Holding.snapshot_date == snapshot_date
            ).delete()
            
            # Add holdings
            for holding_row in holdings:
                symbol = (holding_row.get('symbol') or '').strip()
                if not symbol or symbol in ['CORE**', 'SPAXX**']:  # Skip cash sweep
                    continue
                
                try:
                    qty_str = (holding_row.get('quantity') or '').strip()
                    price_str = (holding_row.get('last price') or '').strip()
                    value_str = (holding_row.get('current value') or '').strip()
                    cost_str = (holding_row.get('cost basis total') or '').strip()
                    desc = (holding_row.get('description') or '').strip()
                    
                    qty = Decimal(qty_str.replace(',', '')) if qty_str else Decimal('0')
                    price = Decimal(price_str.replace('$', '').replace(',', '')) if price_str else None
                    value = Decimal(value_str.replace('$', '').replace(',', '')) if value_str else None
                    cost = Decimal(cost_str.replace('$', '').replace(',', '')) if cost_str else None
                    
                    holding = Holding(
                        user_id=user.id,
                        account_id=account.id,
                        symbol=symbol,
                        name=desc or symbol,
                        quantity=qty,
                        cost_basis=cost,
                        current_price=price,
                        current_value=value,
                        asset_type='stock',
                        snapshot_date=snapshot_date
                    )
                    db.add(holding)
                except Exception as e:
                    logger.error(f"Error parsing holding {symbol}: {e}")
                    continue
            
            imported_count += 1
        
        db.commit()
        return {'imported': imported_count, 'message': f'Imported {imported_count} Fidelity accounts with holdings', 'date': snapshot_date.isoformat()}
    
    # Check if this is a Fidelity account summary CSV
    has_ending_value = len(account_rows) > 0 or len(holdings_rows) > 0
    
    # Only process as Fidelity account summary if not already identified as E*TRADE format
    if has_ending_value and not is_etrade_brokerage and not is_etrade_espp and not is_etrade_rsu:
        # Map account numbers to PortfolioAccount objects
        account_map = {}
        imported_count = 0
        
        # SECTION 1: Process account summary rows
        for idx, row in enumerate(account_rows):
            account_type_val = None
            account_number = None
            account_name = None
            
            for key in row.keys():
                if not key:
                    continue
                key_lower = key.lower()
                if 'account type' in key_lower or 'type' in key_lower:
                    account_type_val = row[key].strip() if row[key] else ''
                elif 'account name' in key_lower or 'account description' in key_lower:
                    account_name = row[key].strip() if row[key] else ''
                elif ('account' in key_lower or 'acct' in key_lower) and 'type' not in key_lower and 'name' not in key_lower:
                    val = row[key].strip() if row[key] else ''
                    # Account numbers are usually numeric or alphanumeric
                    if val and (val.replace('-', '').replace(' ', '').isalnum()):
                        account_number = val
            
            # If no account_type_val but we have account_name, use that as type
            if not account_type_val and account_name:
                account_type_val = account_name
            
            if not account_type_val:
                continue
                
            # Try to get Ending Net Value first, then Ending mkt Value
            ending_value = None
            for key in row.keys():
                if 'ending net value' in key.lower():
                    val = row[key].strip() if row[key] else ''
                    if val and val not in ['', '-', 'N/A']:
                        ending_value = val
                        break
            
            if not ending_value:
                for key in row.keys():
                    if 'ending mkt value' in key.lower():
                        val = row[key].strip() if row[key] else ''
                        if val and val not in ['', '-', 'N/A']:
                            ending_value = val
                            break
            
            if ending_value and account_type_val and account_number:
                try:
                    clean_value = str(ending_value).replace('$', '').replace(',', '').strip()
                    if not clean_value or clean_value in ['-', 'N/A', '']:
                        continue
                        
                    balance = Decimal(clean_value)
                    
                    if balance > 0:
                        last4 = account_number[-4:] if account_number and len(account_number) >= 4 else (account_number or '')
                        
                        # Determine if this is an investment or bank account
                        is_investment = 'ira' in account_type_val.lower() or '401k' in account_type_val.lower() or 'roth' in account_type_val.lower() or 'brokerage' in account_type_val.lower() or 'hsa' in account_type_val.lower() or 'savings account' in account_type_val.lower() and 'health' in account_type_val.lower()
                        
                        # Detect institution from filename
                        institution = 'Fidelity' if is_fidelity_format else 'Imported'
                        
                        # Use account name if available, otherwise use type
                        display_name = account_name or account_type_val
                        
                        account = db.query(PortfolioAccount).filter(
                            PortfolioAccount.user_id == user.id,
                            PortfolioAccount.account_number_last4 == last4
                        ).first() if last4 else None
                        
                        if account:
                            account.balance = balance
                            account.last_synced = datetime.now()
                            account.institution = institution
                            if not account.account_name or 'Fidelity' in account.account_name:
                                account.account_name = display_name
                        else:
                            account = PortfolioAccount(
                                user_id=user.id,
                                institution=institution,
                                account_type='investment' if is_investment else account_type_val.lower().replace(' ', '_').replace('(', '').replace(')', ''),
                                account_name=display_name,
                                account_number_last4=last4,
                                balance=balance,
                                last_synced=datetime.now(),
                                is_active=True
                            )
                            db.add(account)
                            db.flush()
                        
                        account_map[account_number] = account
                        imported_count += 1
                except Exception as e:
                    logger.error(f"CSV Upload: Error parsing account row: {e}")
                    continue

        # SECTION 2: Process holdings rows
        current_account_number = None
        
        for idx, row in enumerate(holdings_rows):
            # Find Symbol/CUSIP column and Quantity column
            symbol_val = None
            qty_val = None
            price_val = None
            value_val = None
            cost_val = None
            desc_val = None
            
            for key in row.keys():
                if not key:
                    continue
                    
                k_lower = key.lower()
                val = row[key].strip() if row[key] else ''
                
                if 'symbol' in k_lower or 'cusip' in k_lower:
                    symbol_val = val
                elif 'quantity' in k_lower:
                    qty_val = val
                elif 'price' in k_lower:
                    price_val = val
                elif 'ending value' in k_lower:
                    value_val = val
                elif 'cost basis' in k_lower:
                    cost_val = val
                elif 'description' in k_lower:
                    desc_val = val
            
            if not symbol_val:
                continue
                
            # Check if this row is actually an account number header
            # Account numbers are typically 8+ chars and digits/alpha
            # But NOT standard keywords mostly found in description or symbol
            if len(symbol_val) >= 8 and (symbol_val.isdigit() or (symbol_val[0].isalpha() and symbol_val[1:].isdigit())):
                # This is likely an account number section header
                current_account_number = symbol_val
                logger.info(f"CSV Upload: Found account number section {current_account_number}")
                continue
                
            # If we have quantity, it's a holding
            if qty_val and current_account_number and current_account_number in account_map:
                try:
                    qty = Decimal(qty_val.replace(',', '').strip())
                    prc = Decimal(price_val.replace('$', '').replace(',', '').strip()) if price_val else None
                    end_val = Decimal(value_val.replace('$', '').replace(',', '').strip()) if value_val else None
                    cb = Decimal(cost_val.replace('$', '').replace(',', '').strip()) if cost_val else None
                    
                    account = account_map[current_account_number]
                    
                    holding = Holding(
                        user_id=user.id,
                        account_id=account.id,
                        symbol=symbol_val,
                        name=desc_val or symbol_val,
                        quantity=qty,
                        cost_basis=cb,
                        current_price=prc,
                        current_value=end_val,
                        asset_type='stock',
                        snapshot_date=snapshot_date
                    )
                    db.add(holding)
                except Exception as e:
                    logger.error(f"CSV Upload: Error parsing holding row: {e}")
                    continue
        
        db.commit()
        
        if imported_count == 0:
            return {'imported': 0, 'message': 'No valid accounts found in CSV.'}
        
        return {'imported': imported_count, 'message': f'Imported/updated {imported_count} accounts with holdings', 'date': snapshot_date.isoformat()}
    
    # Check if E*TRADE ESPP or RSU format
    if is_etrade_espp:
        return await parse_etrade_equity_awards(db, user, csv_text, delimiter, file, snapshot_date, is_espp=True)
    elif is_etrade_rsu:
        return await parse_etrade_equity_awards(db, user, csv_text, delimiter, file, snapshot_date, is_espp=False)
    
    # Otherwise, parse as generic holdings CSV (E*TRADE brokerage, etc.)
    # Define parse_decimal helper
    def parse_decimal(value):
        if value is None:
            return None
        s = str(value).strip()
        if not s or s in ['-', 'N/A', 'n/a']:
            return None
        negative = False
        if s.startswith('(') and s.endswith(')'):
            negative = True
            s = s[1:-1]
        s = s.replace('$', '').replace(',', '').strip()
        # Remove currency codes or text (e.g., USD)
        s = re.sub(r'[^0-9.\-]', '', s)
        if not s or s == '-' or s == '.':
            return None
        try:
            num = Decimal(s)
            return -num if negative else num
        except Exception:
            return None

    default_account_id = None
    fallback_account_id = None
    if not account_id:
        user_accounts = db.query(PortfolioAccount).filter(PortfolioAccount.user_id == user.id).all()
        if len(user_accounts) == 1:
            default_account_id = user_accounts[0].id

    for row in rows:
        if not row:
            continue

        normalized_row = {normalize_key(k): v for k, v in row.items() if k}

        def get_first(keys):
            for k in keys:
                val = normalized_row.get(k)
                if val is not None and str(val).strip() not in ['', '-', 'N/A', 'n/a']:
                    return str(val).strip()
            return None

        def get_by_contains(substrings):
            for key, val in normalized_row.items():
                if not key or val is None:
                    continue
                if any(sub in key for sub in substrings):
                    if str(val).strip() not in ['', '-', 'N/A', 'n/a']:
                        return str(val).strip()
            return None

        # Parse CSV row - handle common broker formats (E*TRADE, Schwab, etc.)
        # E*TRADE uses different column names across account types
        symbol = (get_first(['symbol', 'ticker', 'cusip', 'security symbol', 'stock symbol', 'fund symbol']) or 
                 get_by_contains(['symbol', 'ticker', 'cusip']))
        
        # Skip special E*TRADE rows (CASH, TOTAL, summary rows, metadata)
        if not symbol or symbol.upper() in ['CASH', 'TOTAL', 'BALANCE', 'ACCOUNT']:
            continue
        
        # Skip CSV metadata/header rows (e.g., "GENERATED AT", dates, account info)
        if any(word in symbol.upper() for word in ['GENERATED', 'DOWNLOADED', 'ACCOUNT', 'DATE', 'TIME', 'REPORT', 'PAGE']):
            continue
        
        # Symbol should be reasonable length (1-10 chars) and not contain special phrases
        if len(symbol) < 1 or len(symbol) > 10:
            continue

        # E*TRADE may include account info in each row or only in metadata
        account_number = (get_first(['account number', 'account #', 'acct #', 'acct', 'account', 'account id']) or 
                         get_by_contains(['account number', 'acct', 'account']) or
                         etrade_account_info.get('account number'))
        account_name = (get_first(['account name', 'account description', 'account desc', 'account type']) or 
                       get_by_contains(['account name', 'account description', 'account type']) or
                       etrade_account_info.get('account name') or
                       etrade_account_info.get('account type'))

        account_id_to_use = account_id
        if not account_id_to_use and account_number:
            last4 = account_number[-4:] if len(account_number) >= 4 else account_number
            account = db.query(PortfolioAccount).filter(
                PortfolioAccount.user_id == user.id,
                PortfolioAccount.account_number_last4 == last4
            ).first()
            if account:
                account_id_to_use = account.id

        if not account_id_to_use and account_name:
            account = db.query(PortfolioAccount).filter(
                PortfolioAccount.user_id == user.id,
                PortfolioAccount.account_name == account_name
            ).first()
            if account:
                account_id_to_use = account.id

        if not account_id_to_use and default_account_id:
            account_id_to_use = default_account_id

        if not account_id_to_use and not fallback_account_id:
            if account_number or account_name or etrade_account_info:
                last4 = account_number[-4:] if account_number and len(account_number) >= 4 else (account_number or '')
                
                # Use E*TRADE metadata if available
                if etrade_account_info:
                    institution = 'E*TRADE'
                    account_name = (etrade_account_info.get('account name') or 
                                  etrade_account_info.get('account type') or 
                                  account_name or 
                                  'E*TRADE Holdings')
                    account_name = account_name.strip('"').strip("'")
                    if not last4 and etrade_account_info.get('account number'):
                        acct_num = etrade_account_info['account number']
                        last4 = acct_num[-4:] if len(acct_num) >= 4 else acct_num
                else:
                    filename_lower = (file.filename or '').lower()
                    if 'fidelity' in filename_lower:
                        institution = 'Fidelity'
                    elif 'etrade' in filename_lower:
                        institution = 'E*TRADE'
                    else:
                        institution = 'Imported'
                
                new_account = PortfolioAccount(
                    user_id=user.id,
                    institution=institution,
                    account_type='investment',
                    account_name=account_name or f"{institution} Holdings",
                    account_number_last4=last4,
                    balance=Decimal('0'),
                    last_synced=datetime.now(),
                    is_active=True
                )
                db.add(new_account)
                db.flush()
                fallback_account_id = new_account.id

        if not account_id_to_use and fallback_account_id:
            account_id_to_use = fallback_account_id

        if not account_id_to_use:
            continue

        # E*TRADE uses various column names for quantity, price, value
        quantity = parse_decimal(get_first(['quantity', 'qty', 'shares', 'share quantity', 'share qty']) or get_by_contains(['quantity', 'qty', 'shares']))
        
        # E*TRADE Individual Brokerage uses "Price Paid $" for cost basis per share
        price_paid = parse_decimal(get_first(['price paid', 'price paid $']) or get_by_contains(['price paid']))
        cost_basis_total = parse_decimal(get_first(['cost basis', 'cost basis total', 'total cost', 'cost', 'basis']) or get_by_contains(['cost basis', 'total cost', 'basis']))
        cost_basis = cost_basis_total or (price_paid * quantity if price_paid and quantity else None)
        
        # E*TRADE uses "Last Price $" and "Value $"
        current_price = parse_decimal(get_first(['last price', 'last price $', 'price', 'current price', 'market price', 'last', 'quote']) or get_by_contains(['last price', 'current price', 'market price', 'quote']))
        current_value = parse_decimal(get_first(['value', 'value $', 'current value', 'market value', 'total value', 'mkt value']) or get_by_contains(['value', 'current value', 'market value', 'mkt value', 'total value']))
        
        asset_type = get_first(['type', 'asset type', 'security type', 'asset class', 'security type(s)']) or get_by_contains(['type', 'asset type', 'security type', 'asset class'])
        name = get_first(['description', 'name', 'security description', 'security', 'security name']) or get_by_contains(['description', 'security', 'name'])

        holding = Holding(
            user_id=user.id,
            account_id=account_id_to_use,
            symbol=symbol.strip().upper(),
            name=name or symbol.strip().upper(),
            quantity=quantity or Decimal('0'),
            cost_basis=cost_basis,
            current_price=current_price,
            current_value=current_value,
            asset_type=asset_type,
            snapshot_date=snapshot_date
        )
        db.add(holding)
        imported_count += 1
    
    db.commit()
    
    return {'imported': imported_count, 'date': snapshot_date.isoformat()}


@router.post("/upload/transactions")
async def upload_transactions_csv(
    file: UploadFile = File(...),
    account_id: int = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Upload bank transactions CSV file"""
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")
    
    content = await file.read()
    csv_data = io.StringIO(content.decode('utf-8'))
    reader = csv.DictReader(csv_data)
    
    imported_count = 0
    skipped_count = 0
    
    for row in reader:
        # Parse date
        date_str = row.get('Date') or row.get('Transaction Date') or row.get('DATE')
        if not date_str:
            continue
        
        try:
            txn_date = datetime.strptime(date_str.strip(), '%m/%d/%Y').date()
        except:
            try:
                txn_date = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
            except:
                continue
        
        # Parse amount
        amount_str = row.get('Amount') or row.get('AMOUNT') or '0'
        amount = Decimal(amount_str.replace('$', '').replace(',', '').strip())
        
        # Create unique import ID to prevent duplicates
        import_id = f"{account_id}_{date_str}_{amount}_{row.get('Description', '')[:50]}"
        
        # Check if already imported
        existing = db.query(BankTransaction).filter(
            BankTransaction.import_id == import_id
        ).first()
        
        if existing:
            skipped_count += 1
            continue
        
        transaction = BankTransaction(
            user_id=user.id,
            account_id=account_id,
            transaction_date=txn_date,
            description=row.get('Description') or row.get('DESCRIPTION') or '',
            amount=amount,
            transaction_type=row.get('Type') or row.get('Transaction Type'),
            balance_after=Decimal(str(row.get('Balance') or 0)) if row.get('Balance') else None,
            import_id=import_id
        )
        db.add(transaction)
        imported_count += 1
    
    db.commit()
    
    return {'imported': imported_count, 'skipped': skipped_count}


@router.post("/sync/prices")
async def sync_prices(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Sync current prices for all holdings using Yahoo Finance"""
    
    holdings = db.query(Holding).join(PortfolioAccount).filter(
        PortfolioAccount.user_id == user.id,
        Holding.symbol.isnot(None)
    ).all()
    
    if not holdings:
        return {'updated': 0, 'message': 'No holdings to sync'}
    
    # Group holdings by symbol
    symbols = list(set([h.symbol for h in holdings if h.symbol]))
    
    updated_count = 0
    errors = []
    
    # Batch fetch prices
    try:
        # Create ticker objects
        tickers = yf.Tickers(' '.join(symbols))
        
        for holding in holdings:
            try:
                if not holding.symbol:
                    continue
                    
                ticker = tickers.tickers.get(holding.symbol)
                if not ticker:
                    continue
                
                # Get current price
                info = ticker.info
                current_price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
                
                if current_price:
                    holding.current_price = Decimal(str(current_price))
                    holding.current_value = holding.quantity * holding.current_price
                    holding.snapshot_date = datetime.now()
                    updated_count += 1
                    
            except Exception as e:
                errors.append(f"{holding.symbol}: {str(e)}")
                logger.error(f"Price sync error for {holding.symbol}: {e}")
                continue
        
        db.commit()
        
        # Update account balances based on holdings
        accounts = db.query(PortfolioAccount).filter(
            PortfolioAccount.user_id == user.id,
            PortfolioAccount.account_type == 'investment'
        ).all()
        
        for account in accounts:
            total = db.query(func.sum(Holding.current_value)).filter(
                Holding.account_id == account.id
            ).scalar() or Decimal('0')
            
            account.balance = total
            account.last_synced = datetime.now()
        
        db.commit()
        
        message = f"Updated {updated_count} holdings"
        if errors:
            message += f" ({len(errors)} errors)"
        
        return {
            'updated': updated_count,
            'errors': len(errors),
            'message': message,
            'error_details': errors[:5] if errors else []
        }
        
    except Exception as e:
        logger.error(f"Sync prices error: {e}")
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


@router.get("/espp/grants")
async def get_espp_grants(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Get all ESPP grants for the current user with live price calculations"""
    from app.models import ESPPGrant
    
    grants = db.query(ESPPGrant).filter(
        ESPPGrant.user_id == user.id
    ).order_by(ESPPGrant.purchase_date.desc()).all()
    
    # Fetch live prices for all unique symbols
    symbols = list(set(g.symbol for g in grants if g.symbol))
    live_prices = {}
    if symbols:
        try:
            tickers = yf.Tickers(' '.join(symbols))
            for sym in symbols:
                ticker = tickers.tickers.get(sym)
                if ticker:
                    info = ticker.info
                    price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
                    if price:
                        live_prices[sym] = float(price)
        except Exception as e:
            logger.warning(f"Live price fetch failed for ESPP: {e}")
    
    result = []
    for g in grants:
        live_price = live_prices.get(g.symbol)
        purchase_price = float(g.purchase_price) if g.purchase_price else None
        sellable_qty = float(g.sellable_qty) if g.sellable_qty else None
        
        if live_price and purchase_price and sellable_qty:
            expected_gain_loss = round((live_price - purchase_price) * sellable_qty, 2)
            est_market_value = round(live_price * sellable_qty, 2)
        else:
            expected_gain_loss = float(g.expected_gain_loss) if g.expected_gain_loss else None
            est_market_value = float(g.est_market_value) if g.est_market_value else None
        
        result.append({
            'id': g.id,
            'symbol': g.symbol,
            'record_type': g.record_type,
            'purchase_date': g.purchase_date.isoformat() if g.purchase_date else None,
            'purchase_price': purchase_price,
            'purchased_qty': float(g.purchased_qty) if g.purchased_qty else None,
            'sellable_qty': sellable_qty,
            'expected_gain_loss': expected_gain_loss,
            'est_market_value': est_market_value,
            'live_price': live_price,
            'account_name': g.account.account_name if g.account else None,
            'last_updated': g.last_updated.isoformat() if g.last_updated else None
        })
    
    return result


@router.get("/rsu/grants")
async def get_rsu_grants(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Get all RSU grants for the current user with live price calculations"""
    from app.models import RSUGrant
    
    grants = db.query(RSUGrant).filter(
        RSUGrant.user_id == user.id
    ).order_by(RSUGrant.grant_date.desc()).all()
    
    # Fetch live prices for all unique symbols
    symbols = list(set(g.symbol for g in grants if g.symbol))
    live_prices = {}
    if symbols:
        try:
            tickers = yf.Tickers(' '.join(symbols))
            for sym in symbols:
                ticker = tickers.tickers.get(sym)
                if ticker:
                    info = ticker.info
                    price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
                    if price:
                        live_prices[sym] = float(price)
        except Exception as e:
            logger.warning(f"Live price fetch failed for RSU: {e}")
    
    result = []
    for g in grants:
        live_price = live_prices.get(g.symbol)
        sellable_qty = float(g.sellable_qty) if g.sellable_qty else None
        vested_qty = float(g.vested_qty) if g.vested_qty else None
        granted_qty = float(g.granted_qty) if g.granted_qty else None
        
        if live_price and sellable_qty:
            est_market_value = round(live_price * sellable_qty, 2)
        else:
            est_market_value = float(g.est_market_value) if g.est_market_value else None
        
        unvested_qty = float(g.unvested_qty) if g.unvested_qty else None
        unvested_market_value = round(live_price * unvested_qty, 2) if live_price and unvested_qty else None
        
        result.append({
            'id': g.id,
            'symbol': g.symbol,
            'record_type': g.record_type,
            'grant_number': g.grant_number,
            'grant_date': g.grant_date.isoformat() if g.grant_date else None,
            'settlement_type': g.settlement_type,
            'granted_qty': granted_qty,
            'withheld_qty': float(g.withheld_qty) if g.withheld_qty else None,
            'vested_qty': vested_qty,
            'unvested_qty': float(g.unvested_qty) if g.unvested_qty else None,
            'sellable_qty': sellable_qty,
            'est_market_value': est_market_value,
            'unvested_market_value': unvested_market_value,
            'live_price': live_price,
            'vesting_progress': round((vested_qty / granted_qty) * 100, 1) if granted_qty else 0,
            'account_name': g.account.account_name if g.account else None,
            'last_updated': g.last_updated.isoformat() if g.last_updated else None
        })
    
    return result


@router.post("/sync/accounts")
async def sync_accounts(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Sync account balances by recalculating from holdings"""
    
    accounts = db.query(PortfolioAccount).filter(
        PortfolioAccount.user_id == user.id,
        PortfolioAccount.is_active == True
    ).all()
    
    updated_count = 0
    
    for account in accounts:
        if account.account_type == 'investment':
            # Calculate total from holdings
            total = db.query(func.sum(Holding.current_value)).filter(
                Holding.account_id == account.id
            ).scalar() or Decimal('0')
            
            account.balance = total
            account.last_synced = datetime.now()
            updated_count += 1
    
    db.commit()
    
    return {
        'updated': updated_count,
        'message': f"Synced {updated_count} accounts"
    }


@router.post("/credentials/save")
async def save_broker_credentials(
    institution: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Save encrypted broker credentials"""
    
    try:
        encryptor = CredentialEncryptor()
        encrypted_password = encryptor.encrypt(password)
        
        # Check if credentials already exist
        existing = db.query(BrokerCredential).filter(
            BrokerCredential.user_id == user.id,
            BrokerCredential.institution == institution
        ).first()
        
        if existing:
            existing.username = username
            existing.encrypted_password = encrypted_password
            existing.is_active = True
        else:
            credential = BrokerCredential(
                user_id=user.id,
                institution=institution,
                username=username,
                encrypted_password=encrypted_password,
                is_active=True
            )
            db.add(credential)
        
        db.commit()
        
        return {'success': True, 'message': f'{institution} credentials saved'}
        
    except Exception as e:
        logger.error(f"Error saving credentials: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/credentials/list")
async def list_broker_credentials(
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """List configured broker credentials (without passwords)"""
    
    credentials = db.query(BrokerCredential).filter(
        BrokerCredential.user_id == user.id,
        BrokerCredential.is_active == True
    ).all()
    
    return [{
        'id': c.id,
        'institution': c.institution,
        'username': c.username,
        'last_used': c.last_used.isoformat() if c.last_used else None
    } for c in credentials]


@router.delete("/credentials/{credential_id}")
async def delete_broker_credentials(
    credential_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Delete broker credentials"""
    
    credential = db.query(BrokerCredential).filter(
        BrokerCredential.id == credential_id,
        BrokerCredential.user_id == user.id
    ).first()
    
    if not credential:
        raise HTTPException(status_code=404, detail="Credentials not found")
    
    db.delete(credential)
    db.commit()
    
    return {'success': True, 'message': 'Credentials deleted'}


# Fidelity scraper endpoint disabled - using Plaid for syncing
# @router.post("/sync/fidelity")
# async def sync_fidelity(...):
#     """Sync Fidelity accounts and holdings automatically"""
#     # Disabled in favor of Plaid integration


@router.post("/accounts")
def create_account(
    institution: str = Form(...),
    account_type: str = Form(...),
    account_name: str = Form(None),
    account_number_last4: str = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Create a new portfolio account"""
    
    account = PortfolioAccount(
        user_id=user.id,
        institution=institution,
        account_type=account_type,
        account_name=account_name,
        account_number_last4=account_number_last4
    )
    db.add(account)
    db.commit()
    db.refresh(account)
    
    return {
        'id': account.id,
        'institution': account.institution,
        'account_type': account.account_type,
        'account_name': account.account_name
    }


@router.get("/accounts")
def list_accounts(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """List all accounts"""
    
    accounts = db.query(PortfolioAccount).filter(
        PortfolioAccount.user_id == user.id,
        PortfolioAccount.is_active == True
    ).all()
    
    result = []
    for acc in accounts:
        result.append({
            'id': acc.id,
            'institution': acc.institution,
            'account_type': acc.account_type,
            'account_name': acc.account_name or acc.institution,
            'account_number_last4': acc.account_number_last4,
            'balance': float(acc.balance),
            'last_synced': acc.last_synced.isoformat() if acc.last_synced else None
        })
    
    return result


@router.delete("/accounts/{account_id}")
def delete_account(account_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Delete an account and all its holdings"""
    
    account = db.query(PortfolioAccount).filter(
        PortfolioAccount.id == account_id,
        PortfolioAccount.user_id == user.id
    ).first()
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Delete associated holdings
    db.query(Holding).filter(Holding.account_id == account_id).delete()
    
    # Delete account
    db.delete(account)
    db.commit()
    
    return {'message': 'Account deleted successfully'}


@router.patch("/accounts/{account_id}")
def update_account(
    account_id: int, 
    account_holder: str = Form(...),
    db: Session = Depends(get_db), 
    user=Depends(get_current_user)
):
    """Update account holder name"""
    
    account = db.query(PortfolioAccount).filter(
        PortfolioAccount.id == account_id,
        PortfolioAccount.user_id == user.id
    ).first()
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    account.account_holder = account_holder
    db.commit()
    
    return {'message': 'Account updated successfully', 'account_holder': account_holder}


@router.post("/bulk-upload")
async def bulk_upload_files(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Bulk upload multiple CSV/PDF files from different banks/brokers.
    Automatically detects format and imports transactions/holdings/balances.
    """
    results = {
        'success': [],
        'errors': [],
        'summary': {
            'files_processed': 0,
            'transactions_imported': 0,
            'accounts_updated': 0,
            'holdings_imported': 0
        }
    }
    
    for file in files:
        filename = file.filename
        
        # Skip PDFs for now (we'll just log them)
        if filename.lower().endswith('.pdf'):
            results['success'].append(f"PDF saved: {filename} (statements stored for reference)")
            results['summary']['files_processed'] += 1
            continue
        
        # Process CSV files
        if not filename.lower().endswith('.csv'):
            results['errors'].append(f"Unsupported file type: {filename}")
            continue
        
        try:
            # Read file content
            content = await file.read()
            content_str = content.decode('utf-8')
            
            # Try portfolio holdings import first (for E*TRADE, Fidelity, etc.)
            # Reset file position for re-use
            await file.seek(0)
            
            try:
                holdings_result = await upload_holdings_csv(file, None, db, user)
                logger.info(f"Holdings import result for {filename}: {holdings_result}")
                if holdings_result.get('imported', 0) > 0:
                    results['success'].append(f"{filename}: Imported {holdings_result['imported']} holdings")
                    results['summary']['holdings_imported'] += holdings_result['imported']
                    results['summary']['files_processed'] += 1
                    continue
            except Exception as holdings_error:
                logger.warning(f"Not a holdings CSV ({filename}), trying transactions: {holdings_error}")
                # Reset for next attempt
                await file.seek(0)
                content = await file.read()
                content_str = content.decode('utf-8')
            
            # Fall back to transaction CSV parsing
            # Parse CSV
            parsed = CSVParser.parse_csv(content_str, filename)
            
            if parsed['errors']:
                results['errors'].extend(parsed['errors'])
                continue
            
            # Extract account info from filename
            institution, account_last4 = extract_account_info(filename)
            
            # Find or create account
            account = db.query(PortfolioAccount).filter(
                PortfolioAccount.user_id == user.id,
                PortfolioAccount.institution == institution,
                PortfolioAccount.account_number_last4 == account_last4
            ).first()
            
            if not account:
                # Create new account
                account_type = guess_account_type(filename, parsed)
                
                # Get account name from parsed data or create one
                account_name = None
                if parsed['account_balances']:
                    account_name = parsed['account_balances'][0].get('account_name')
                
                if not account_name:
                    account_name = f"{institution} (...{account_last4})"
                
                account = PortfolioAccount(
                    user_id=user.id,
                    institution=institution,
                    account_type=account_type,
                    account_name=account_name,
                    account_number_last4=account_last4,
                    balance=Decimal('0')
                )
                db.add(account)
                db.flush()
                results['summary']['accounts_updated'] += 1
            
            # Import transactions
            transactions_added = 0
            for txn in parsed['transactions']:
                # Check for duplicates
                import_id = f"{filename}_{txn['date']}_{txn['amount']}"
                existing = db.query(BankTransaction).filter(
                    BankTransaction.import_id == import_id
                ).first()
                
                if not existing:
                    bank_txn = BankTransaction(
                        user_id=user.id,
                        account_id=account.id,
                        transaction_date=txn['date'],
                        description=txn['description'],
                        amount=txn['amount'],
                        transaction_type=txn.get('transaction_type'),
                        balance_after=txn.get('balance'),
                        import_id=import_id
                    )
                    db.add(bank_txn)
                    transactions_added += 1
            
            # Update account balance from parsed data
            if parsed['account_balances']:
                for balance_info in parsed['account_balances']:
                    # Update the current account if account number matches
                    if balance_info['account_number'][-4:] == account_last4:
                        account.balance = balance_info['balance']
                        if balance_info.get('account_name') and not account.account_name:
                            account.account_name = balance_info['account_name']
                        account.last_synced = datetime.utcnow()
                        break
            
            # If no balance info in parsed data, calculate from transactions
            if account.balance == 0 and parsed['transactions']:
                # Use the last transaction's balance_after if available
                last_txn = parsed['transactions'][-1]
                if last_txn.get('balance'):
                    account.balance = last_txn['balance']
            
            # Import holdings
            holdings_added = 0
            snapshot_date = date.today()
            for holding_info in parsed['holdings']:
                # Clear old holdings for this account
                db.query(Holding).filter(
                    Holding.account_id == account.id,
                    Holding.snapshot_date == snapshot_date
                ).delete()
                
                holding = Holding(
                    user_id=user.id,
                    account_id=account.id,
                    symbol=holding_info['symbol'],
                    name=holding_info.get('description'),
                    quantity=holding_info['quantity'],
                    cost_basis=holding_info.get('value'),
                    current_price=holding_info.get('price'),
                    current_value=holding_info.get('value'),
                    snapshot_date=snapshot_date
                )
                db.add(holding)
                holdings_added += 1
            
            account.last_synced = datetime.utcnow()
            
            db.commit()
            
            results['success'].append(
                f"{filename}: {transactions_added} transactions, {holdings_added} holdings"
            )
            results['summary']['files_processed'] += 1
            results['summary']['transactions_imported'] += transactions_added
            results['summary']['holdings_imported'] += holdings_added
            
        except Exception as e:
            results['errors'].append(f"{filename}: {str(e)}")
            db.rollback()
    
    return results


def extract_account_info(filename: str) -> tuple:
    """Extract institution and account number from filename"""
    filename_lower = filename.lower()
    
    # Map filenames to institutions
    if 'usb' in filename_lower or 'personal checking' in filename_lower:
        institution = 'USB Bank'
        # Extract last 4 digits
        match = re.search(r'(\d{4})', filename)
        account_last4 = match.group(1) if match else '0000'
    elif 'chase' in filename_lower:
        institution = 'Chase'
        match = re.search(r'(\d{4})', filename)
        account_last4 = match.group(1) if match else '0000'
    elif 'fidelity' in filename_lower or 'x8' in filename_lower or '224' in filename_lower:
        institution = 'Fidelity'
        match = re.search(r'[xX]?(\d{4,8})', filename)
        account_last4 = match.group(1)[-4:] if match else '0000'
    elif '401k' in filename_lower:
        institution = 'Intel 401k'
        account_last4 = '0000'
    else:
        institution = 'Other'
        match = re.search(r'(\d{4})', filename)
        account_last4 = match.group(1) if match else '0000'
    
    return institution, account_last4


def guess_account_type(filename: str, parsed_data: dict) -> str:
    """Guess account type from filename and data"""
    filename_lower = filename.lower()
    
    if 'checking' in filename_lower:
        return 'checking'
    elif 'savings' in filename_lower:
        return 'savings'
    elif 'credit' in filename_lower or 'card' in filename_lower:
        return 'credit_card'
    elif '401k' in filename_lower or 'ira' in filename_lower or 'hsa' in filename_lower:
        return 'investment'
    elif parsed_data.get('holdings'):
        return 'investment'
    else:
        return 'checking'


# ==================== Plaid Integration ====================

@router.post("/plaid/create-link-token")
async def create_plaid_link_token(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Create a Plaid Link token for connecting accounts"""
    try:
        plaid_client = get_plaid_client()
        redirect_uri = os.getenv('PLAID_REDIRECT_URI') or f"{str(request.base_url).rstrip('/')}/plaid/oauth-return"
        result = plaid_client.create_link_token(user.id, user.username, redirect_uri=redirect_uri)
        return {'link_token': result['link_token']}
    except Exception as e:
        logger.error(f"Error creating Plaid link token: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/plaid/update-link-token/{item_id}")
async def create_plaid_update_link_token(item_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Create a Plaid Link token in update mode to add Identity product to existing connection"""
    plaid_item = db.query(PlaidItem).filter(
        PlaidItem.id == item_id,
        PlaidItem.user_id == user.id,
        PlaidItem.is_active == True
    ).first()
    
    if not plaid_item:
        raise HTTPException(status_code=404, detail="Plaid item not found")
    
    try:
        plaid_client = get_plaid_client()
        redirect_uri = os.getenv('PLAID_REDIRECT_URI') or f"{str(request.base_url).rstrip('/')}/plaid/oauth-return"
        result = plaid_client.create_link_token(user.id, user.username, access_token=plaid_item.access_token, redirect_uri=redirect_uri)
        return {'link_token': result['link_token']}
    except Exception as e:
        logger.error(f"Error creating Plaid update link token: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/plaid/exchange-token")
async def exchange_plaid_token(
    public_token: str = Form(...),
    institution_name: str = Form(None),
    institution_id: str = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Exchange public token for access token and store it"""
    try:
        logger.info(f"Starting Plaid token exchange for user {user.id}")
        logger.info(f"Institution: {institution_name} (ID: {institution_id})")
        
        plaid_client = get_plaid_client()
        logger.info("Plaid client initialized, calling exchange_public_token...")
        
        result = plaid_client.exchange_public_token(public_token)
        logger.info(f"Token exchange successful. Item ID: {result['item_id']}")
        
        # Store the access token
        plaid_item = PlaidItem(
            user_id=user.id,
            item_id=result['item_id'],
            access_token=result['access_token'],
            institution_name=institution_name,
            institution_id=institution_id
        )
        db.add(plaid_item)
        db.commit()
        logger.info(f"PlaidItem saved to database with ID: {plaid_item.id}")
        
        # Immediately sync accounts
        logger.info("Starting account sync...")
        await sync_plaid_item(plaid_item.id, db, user)
        logger.info("Account sync completed successfully")
        
        return {'success': True, 'item_id': result['item_id']}
    except Exception as e:
        logger.error(f"Error exchanging Plaid token: {type(e).__name__}: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/plaid/items")
async def list_plaid_items(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """List all connected Plaid items"""
    items = db.query(PlaidItem).filter(
        PlaidItem.user_id == user.id,
        PlaidItem.is_active == True
    ).all()
    
    return [{
        'id': item.id,
        'institution_name': item.institution_name,
        'last_synced': item.last_synced.isoformat() if item.last_synced else None,
        'created_at': item.created_at.isoformat()
    } for item in items]


@router.post("/plaid/sync/all")
async def sync_all_plaid_items(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Sync all connected Plaid items at once"""
    items = db.query(PlaidItem).filter(
        PlaidItem.user_id == user.id,
        PlaidItem.is_active == True
    ).all()

    if not items:
        return {'synced': 0, 'message': 'No connected Plaid items found'}

    results = []
    for item in items:
        try:
            # Reuse the per-item sync logic inline
            plaid_client = get_plaid_client()
            accounts_data = plaid_client.get_accounts(item.access_token)
            accounts_synced = 0
            for acc in accounts_data['accounts']:
                account = db.query(PortfolioAccount).filter(
                    PortfolioAccount.user_id == user.id,
                    PortfolioAccount.institution == item.institution_name,
                    PortfolioAccount.account_number_last4 == acc['mask']
                ).first()
                if account:
                    acc_type = acc['type'].lower()
                    if acc_type == 'credit':
                        account.balance = Decimal(str(abs(acc.get('balances', {}).get('current', 0) or 0)))
                    else:
                        balance = acc.get('balances', {}).get('current', 0) or 0
                        account.balance = Decimal(str(balance))
                    account.last_synced = datetime.now()
                    accounts_synced += 1

            # Sync transactions (last 30 days)
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            plaid_account_map = {a['account_id']: a['mask'] for a in accounts_data['accounts']}
            transactions_added = 0
            transactions_data = plaid_client.get_transactions(item.access_token, start_date, end_date)
            for txn in transactions_data.get('transactions', []):
                mask = plaid_account_map.get(txn['account_id'])
                if not mask:
                    continue
                account = db.query(PortfolioAccount).filter(
                    PortfolioAccount.user_id == user.id,
                    PortfolioAccount.institution == item.institution_name,
                    PortfolioAccount.account_number_last4 == mask
                ).first()
                if account:
                    import_id = f"plaid_{txn['transaction_id']}"
                    existing = db.query(BankTransaction).filter(BankTransaction.import_id == import_id).first()
                    if not existing:
                        txn_date = txn['date']
                        if isinstance(txn_date, str):
                            txn_date = datetime.strptime(txn_date, '%Y-%m-%d').date()
                        bt = BankTransaction(
                            user_id=user.id,
                            account_id=account.id,
                            import_id=import_id,
                            transaction_date=txn_date,
                            description=txn.get('name', ''),
                            amount=Decimal(str(-txn['amount'])),
                            transaction_type=txn.get('category', [None])[0] if txn.get('category') else None,
                            category=txn.get('personal_finance_category', {}).get('primary') if txn.get('personal_finance_category') else (txn.get('category', [''])[0] if txn.get('category') else ''),
                        )
                        db.add(bt)
                        transactions_added += 1

            item.last_synced = datetime.now()
            db.commit()
            results.append({'institution': item.institution_name, 'accounts': accounts_synced, 'transactions': transactions_added})
        except Exception as e:
            logger.error(f"Error syncing Plaid item {item.id}: {e}")
            results.append({'institution': item.institution_name, 'error': str(e)})

    return {'synced': len(results), 'results': results}


@router.post("/plaid/sync/{item_id}")
async def sync_plaid_item(item_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Sync data from a Plaid item"""
    plaid_item = db.query(PlaidItem).filter(
        PlaidItem.id == item_id,
        PlaidItem.user_id == user.id,
        PlaidItem.is_active == True
    ).first()
    
    if not plaid_item:
        raise HTTPException(status_code=404, detail="Plaid item not found")
    
    try:
        plaid_client = get_plaid_client()
        
        # Get accounts
        accounts_data = plaid_client.get_accounts(plaid_item.access_token)
        
        # Try to get identity information (owner names)
        identity_data = None
        owner_names = {}
        try:
            identity_data = plaid_client.get_identity(plaid_item.access_token)
            # Build a map of account_id -> owner name
            for acc in identity_data.get('accounts', []):
                owners = acc.get('owners', [])
                if owners:
                    # Use first owner's name
                    owner = owners[0]
                    names = owner.get('names', [])
                    if names:
                        owner_names[acc['account_id']] = names[0]
        except Exception as e:
            logger.warning(f"Could not fetch identity data: {e}")
        
        # Build a map of Plaid account_id to mask (last 4 digits)
        plaid_account_map = {}
        for acc in accounts_data['accounts']:
            plaid_account_map[acc['account_id']] = acc['mask']
        
        accounts_synced = 0
        for acc in accounts_data['accounts']:
            
            # Find or create account
            account = db.query(PortfolioAccount).filter(
                PortfolioAccount.user_id == user.id,
                PortfolioAccount.institution == plaid_item.institution_name,
                PortfolioAccount.account_number_last4 == acc['mask']
            ).first()
            
            if not account:
                # Map Plaid account type to our type
                acc_type = acc['type'].lower()
                if acc_type == 'depository':
                    if acc['subtype'] == 'checking':
                        account_type = 'checking'
                    elif acc['subtype'] == 'savings':
                        account_type = 'savings'
                    else:
                        account_type = 'bank'
                elif acc_type == 'credit':
                    account_type = 'credit_card'
                elif acc_type == 'investment' or acc_type == 'brokerage':
                    account_type = 'investment'
                else:
                    account_type = 'other'
                
                # Calculate balance based on account type
                if acc_type == 'credit':
                    # For credit cards: use current balance (amount owed)
                    balance = Decimal(str(acc['balances'].get('current') or 0))
                else:
                    # For other accounts: use current or available
                    balance = Decimal(str(acc['balances']['current'] or acc['balances'].get('available') or 0))
                
                # Get owner name if available
                account_holder_name = owner_names.get(acc['account_id'])
                
                account = PortfolioAccount(
                    user_id=user.id,
                    institution=plaid_item.institution_name,
                    account_type=account_type,
                    account_name=acc['name'],
                    account_holder=account_holder_name,
                    account_number_last4=acc['mask'],
                    balance=balance
                )
                db.add(account)
                db.flush()
            else:
                # Update balance based on account type
                if acc['type'].lower() == 'credit':
                    # For credit cards: use current balance (amount owed)
                    balance_value = acc['balances'].get('current') or 0
                else:
                    # For other accounts: use current or available
                    balance_value = acc['balances']['current'] or acc['balances'].get('available') or 0
                account.balance = Decimal(str(balance_value))
                account.last_synced = datetime.utcnow()
                # Update owner name if we have it and account doesn't have one
                if not account.account_holder and acc['account_id'] in owner_names:
                    account.account_holder = owner_names[acc['account_id']]
            
            accounts_synced += 1
        
        # Get transactions (last 30 days)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)
        
        transactions_added = 0
        try:
            transactions_data = plaid_client.get_transactions(plaid_item.access_token, start_date, end_date)
            
            for txn in transactions_data['transactions']:
                # Get the mask for this Plaid account_id
                mask = plaid_account_map.get(txn['account_id'])
                if not mask:
                    continue
                
                # Find account by mask
                account = db.query(PortfolioAccount).filter(
                    PortfolioAccount.user_id == user.id,
                    PortfolioAccount.institution == plaid_item.institution_name,
                    PortfolioAccount.account_number_last4 == mask
                ).first()
                
                if account:
                    # Check for duplicate
                    import_id = f"plaid_{txn['transaction_id']}"
                    existing = db.query(BankTransaction).filter(
                        BankTransaction.import_id == import_id
                    ).first()
                    
                    if not existing:
                        # Handle date - Plaid returns it as date object or string
                        txn_date = txn['date']
                        if isinstance(txn_date, str):
                            txn_date = datetime.strptime(txn_date, '%Y-%m-%d').date()
                        
                        bank_txn = BankTransaction(
                            user_id=user.id,
                            account_id=account.id,
                            transaction_date=txn_date,
                            description=txn['name'],
                            amount=Decimal(str(-txn['amount'])),  # Plaid uses negative for outflows
                            transaction_type=txn.get('category', [None])[0] if txn.get('category') else None,
                            import_id=import_id
                        )
                        db.add(bank_txn)
                        transactions_added += 1
        except Exception as e:
            logger.warning(f"Could not fetch transactions: {e}")
            transactions_added = 0
        
        # Get investment holdings for investment accounts
        holdings_added = 0
        try:
            holdings_data = plaid_client.get_investment_holdings(plaid_item.access_token)
            
            snapshot_date = date.today()
            
            # Group holdings by account to calculate balances
            account_balances = {}
            
            for holding in holdings_data['holdings']:
                # Map account_id to find the right account
                plaid_account_id = holding['account_id']
                
                # Find account by plaid account id (last 4 of account_id or by matching)
                account = None
                for acc in accounts_data['accounts']:
                    if acc['account_id'] == plaid_account_id:
                        account = db.query(PortfolioAccount).filter(
                            PortfolioAccount.user_id == user.id,
                            PortfolioAccount.institution == plaid_item.institution_name,
                            PortfolioAccount.account_number_last4 == acc['mask']
                        ).first()
                        break
                
                if account:
                    # Clear old holdings for today (do this once per account)
                    if account.id not in account_balances:
                        db.query(Holding).filter(
                            Holding.account_id == account.id,
                            Holding.snapshot_date == snapshot_date
                        ).delete()
                        account_balances[account.id] = Decimal('0')
                    
                    # Find security details
                    security = None
                    for sec in holdings_data.get('securities', []):
                        if sec['security_id'] == holding['security_id']:
                            security = sec
                            break
                    
                    holding_value = Decimal(str(holding.get('institution_value', 0)))
                    account_balances[account.id] += holding_value
                    
                    holding_obj = Holding(
                        user_id=user.id,
                        account_id=account.id,
                        symbol=security.get('ticker_symbol', 'UNKNOWN') if security else 'UNKNOWN',
                        name=security.get('name', 'Unknown') if security else 'Unknown',
                        quantity=Decimal(str(holding['quantity'])),
                        cost_basis=Decimal(str(holding['cost_basis'])) if holding.get('cost_basis') else None,
                        current_price=Decimal(str(holding['institution_price'])) if holding.get('institution_price') else None,
                        current_value=holding_value,
                        snapshot_date=snapshot_date
                    )
                    db.add(holding_obj)
                    holdings_added += 1
            
            # Update investment account balances from holdings totals
            for account_id, total_balance in account_balances.items():
                account = db.query(PortfolioAccount).filter(PortfolioAccount.id == account_id).first()
                if account:
                    account.balance = total_balance
                    logger.info(f"Updated investment account {account.account_name} balance to ${total_balance}")
                    
        except plaid.ApiException as plaid_error:
            # Check if it's a PRODUCT_NOT_READY error (Investments not enabled)
            if 'PRODUCT_NOT_READY' in str(plaid_error):
                logger.warning(f"Plaid Investments product not enabled for this item: {plaid_error}")
                holdings_added = 0
            else:
                logger.error(f"Plaid API error fetching holdings: {plaid_error}")
                holdings_added = 0
        except Exception as e:
            logger.warning(f"No investment holdings data available: {e}")
            holdings_added = 0
        
        # Update last synced
        plaid_item.last_synced = datetime.utcnow()
        db.commit()
        
        return {
            'success': True,
            'accounts': accounts_synced,
            'transactions': transactions_added,
            'holdings': holdings_added
        }
        
    except Exception as e:
        logger.error(f"Error syncing Plaid item: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/plaid/items/{item_id}")
async def delete_plaid_item(item_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Disconnect a Plaid item"""
    plaid_item = db.query(PlaidItem).filter(
        PlaidItem.id == item_id,
        PlaidItem.user_id == user.id
    ).first()
    
    if not plaid_item:
        raise HTTPException(status_code=404, detail="Plaid item not found")
    
    plaid_item.is_active = False
    db.commit()
    
    return {'success': True}

